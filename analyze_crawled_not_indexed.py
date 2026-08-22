from pathlib import Path
import csv,re,subprocess
from openpyxl import load_workbook

xlsx=Path('/home/ubuntu/upload/https___www.canticosccb.com.br_-Coverage-Drilldown-2026-08-21.xlsx')
wb=load_workbook(xlsx,read_only=True,data_only=True)
ws=wb['Tabela']; rows=list(ws.iter_rows(values_only=True)); headers=[str(x or '').strip() for x in rows[0]]
ui=headers.index('URL'); li=headers.index('Último rastreamento')
urls=[]
for r in rows[1:]:
    if not r[ui]: continue
    url=str(r[ui]).strip().replace('\n',' ')
    urls.append((url,str(r[li] or '').strip()))
sitemap=Path('/home/ubuntu/canticosccb-fix/public/sitemap.xml').read_text(errors='ignore')
sitemap_urls=set(re.findall(r'<loc>([^<]+)</loc>',sitemap))
fields=['url','last_crawl','family','in_sitemap','http_status','final_url','robots','canonical','title','h1','body_bytes','content_signal','classification']
out=Path('/home/ubuntu/canticosccb-fix/CRAWLED_NOT_INDEXED_VALIDATION.csv')
def clean(x): return re.sub(r'\s+',' ',re.sub('<[^>]+>',' ',x or '')).strip()
def extract(body,pat):
    x=re.search(pat,body,re.I|re.S); return clean(x.group(1)) if x else ''
def family(url):
    p=re.sub(r'^https?://[^/]+','',url).split('?',1)[0]
    return '/' + p.strip('/').split('/')[0] if p.strip('/') else '/'
with out.open('w',newline='',encoding='utf-8') as f:
    w=csv.DictWriter(f,fieldnames=fields); w.writeheader()
    for url,last in urls:
        hp='/tmp/cni-head'; bp='/tmp/cni-body'
        p=subprocess.run(['curl','-L','-sS','-A','Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)','-D',hp,'-o',bp,'-w','STATUS:%{http_code}\nFINAL:%{url_effective}\n','--max-time','25',url],capture_output=True,text=True)
        body=Path(bp).read_text(errors='ignore') if Path(bp).exists() else ''
        status=re.findall(r'STATUS:(\d+)',p.stdout); final=re.findall(r'FINAL:(.*)',p.stdout)
        robots=extract(body,r'<meta[^>]+name=["\']robots["\'][^>]+content=["\']([^"\']+)')
        canonical=extract(body,r'<link[^>]+rel=["\']canonical["\'][^>]+href=["\']([^"\']+)')
        title=extract(body,r'<title[^>]*>(.*?)</title>'); h1=extract(body,r'<h1[^>]*>(.*?)</h1>')
        st=int(status[-1]) if status else None
        signal=any(s in body.lower() for s in ['serviço temporariamente indisponível','nenhum hino','nenhuma cifra','não encontrado','conteúdo indisponível'])
        if st in (404,410): cls='real_404'
        elif st and st>=500: cls='server_error'
        elif st==200 and robots and 'noindex' in robots.lower(): cls='noindex_200'
        elif st==200 and not h1 and len(body)<5000: cls='thin_200'
        elif st==200 and signal: cls='content_signal'
        elif st==200: cls='200_indexable_candidate'
        elif st and 300<=st<400: cls='redirect'
        else: cls='request_failed'
        w.writerow({'url':url,'last_crawl':last,'family':family(url),'in_sitemap':url in sitemap_urls,'http_status':st or '', 'final_url':final[-1].strip() if final else '', 'robots':robots,'canonical':canonical,'title':title,'h1':h1,'body_bytes':len(body.encode()),'content_signal':signal,'classification':cls})
        print(st,cls,family(url),url)
print(f'WROTE {out} ({len(urls)} URLs)')
