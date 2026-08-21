from pathlib import Path
import csv, re, subprocess
from openpyxl import load_workbook

xlsx = Path('/home/ubuntu/upload/https___www.canticosccb.com.br_-Coverage-Drilldown-2026-08-20(1).xlsx')
wb = load_workbook(xlsx, read_only=True, data_only=True)
ws = wb['Tabela']
rows = list(ws.iter_rows(values_only=True))
headers = [str(x or '').strip() for x in rows[0]]
url_i = headers.index('URL')
last_i = headers.index('Último rastreamento')
urls = [(str(r[url_i]).strip(), str(r[last_i] or '').strip()) for r in rows[1:] if r[url_i]]
sitemap = Path('/home/ubuntu/canticosccb-fix/public/sitemap.xml').read_text(errors='ignore')
sitemap_urls = set(re.findall(r'<loc>([^<]+)</loc>', sitemap))
fields=['url','last_crawl','in_sitemap','http_status','final_url','content_type','robots','canonical','title','h1','body_bytes','soft404_signal','classification']
out=Path('/home/ubuntu/canticosccb-fix/SOFT404_URLS_VALIDATION.csv')
with out.open('w',newline='',encoding='utf-8') as f:
    w=csv.DictWriter(f,fieldnames=fields); w.writeheader()
    for url,last in urls:
        hp='/tmp/soft-head'; bp='/tmp/soft-body'
        p=subprocess.run(['curl','-L','-sS','-A','Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)','-D',hp,'-o',bp,'-w','STATUS:%{http_code}\nFINAL:%{url_effective}\nTYPE:%{content_type}\n','--max-time','25',url],capture_output=True,text=True)
        body=Path(bp).read_text(errors='ignore') if Path(bp).exists() else ''
        status=re.findall(r'STATUS:(\d+)',p.stdout); final=re.findall(r'FINAL:(.*)',p.stdout); ctype=re.findall(r'TYPE:(.*)',p.stdout)
        def m(pat):
            x=re.search(pat,body,re.I|re.S); return re.sub(r'\s+',' ',re.sub('<[^>]+>',' ',x.group(1) if x else '')).strip() if x else ''
        robots=m(r'<meta[^>]+name=["\']robots["\'][^>]+content=["\']([^"\']+)')
        canonical=m(r'<link[^>]+rel=["\']canonical["\'][^>]+href=["\']([^"\']+)')
        title=m(r'<title[^>]*>(.*?)</title>'); h1=m(r'<h1[^>]*>(.*?)</h1>')
        st=int(status[-1]) if status else None
        soft='não encontrado' in body.lower() or 'nenhum hino' in body.lower() or 'serviço temporariamente' in body.lower() or st in (404,410)
        if st in (404,410): cls='real_404'
        elif st and st>=500: cls='server_error'
        elif st and 300<=st<400: cls='redirect'
        elif st==200 and soft: cls='200_soft_signal'
        elif st==200: cls='200_content'
        else: cls='request_failed'
        w.writerow({'url':url,'last_crawl':last,'in_sitemap':url in sitemap_urls,'http_status':st or '', 'final_url':final[-1].strip() if final else '', 'content_type':ctype[-1].strip() if ctype else '', 'robots':robots,'canonical':canonical,'title':title,'h1':h1,'body_bytes':len(body.encode()),'soft404_signal':soft,'classification':cls})
        print(st,cls,url)
print(f'WROTE {out} ({len(urls)} URLs)')
