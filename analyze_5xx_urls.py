from pathlib import Path
from urllib.parse import urlparse
import csv, re, subprocess
from openpyxl import load_workbook

xlsx = Path('/home/ubuntu/upload/https___www.canticosccb.com.br_-Coverage-Drilldown-2026-08-20.xlsx')
wb = load_workbook(xlsx, read_only=True, data_only=True)
ws = wb['Tabela']
rows = list(ws.iter_rows(values_only=True))
headers = [str(x or '').strip() for x in rows[0]]
url_i = headers.index('URL')
last_i = headers.index('Último rastreamento')
urls = [(str(r[url_i]).strip(), str(r[last_i] or '').strip()) for r in rows[1:] if r[url_i]]

sitemap = Path('/home/ubuntu/canticosccb-fix/public/sitemap.xml').read_text(errors='ignore')
sitemap_urls = set(re.findall(r'<loc>([^<]+)</loc>', sitemap))

out = Path('/home/ubuntu/canticosccb-fix/5XX_URLS_VALIDATION.csv')
fields = ['url','last_crawl','in_sitemap','http_status','final_url','content_type','robots','canonical','title','h1','body_bytes','server_503_signal','classification']
with out.open('w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=fields)
    writer.writeheader()
    for url, last in urls:
        head_path='/tmp/5xx-head'; body_path='/tmp/5xx-body'
        cmd=['curl','-L','-sS','-A','Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)','-D',head_path,'-o',body_path,'-w','STATUS:%{http_code}\nFINAL:%{url_effective}\nTYPE:%{content_type}\n','--max-time','25',url]
        proc=subprocess.run(cmd, capture_output=True, text=True)
        head=Path(head_path).read_text(errors='ignore') if Path(head_path).exists() else ''
        body=Path(body_path).read_text(errors='ignore') if Path(body_path).exists() else ''
        status=re.findall(r'STATUS:(\d+)',proc.stdout)
        final=re.findall(r'FINAL:(.*)',proc.stdout)
        ctype=re.findall(r'TYPE:(.*)',proc.stdout)
        robots=re.search(r'<meta[^>]+name=["\']robots["\'][^>]+content=["\']([^"\']+)',body,re.I)
        canonical=re.search(r'<link[^>]+rel=["\']canonical["\'][^>]+href=["\']([^"\']+)',body,re.I)
        title=re.search(r'<title[^>]*>(.*?)</title>',body,re.I|re.S)
        h1=re.search(r'<h1[^>]*>(.*?)</h1>',body,re.I|re.S)
        clean=lambda x: re.sub(r'\s+',' ',re.sub('<[^>]+>',' ',x or '')).strip()
        st=int(status[-1]) if status else None
        signal=any(x in body.lower() for x in ['serviço temporariamente indisponível','supabase indisponível','erro no servidor','temporarily unavailable'])
        if st and st >= 500: cls='still_5xx'
        elif st == 404: cls='now_404'
        elif st and 300 <= st < 400: cls='redirect'
        elif st == 200 and signal: cls='200_error_body'
        elif st == 200: cls='recovered_200'
        else: cls='request_failed'
        writer.writerow({'url':url,'last_crawl':last,'in_sitemap':url in sitemap_urls,'http_status':st or '', 'final_url':final[-1].strip() if final else '', 'content_type':ctype[-1].strip() if ctype else '', 'robots':robots.group(1).strip() if robots else '', 'canonical':canonical.group(1).strip() if canonical else '', 'title':clean(title.group(1) if title else ''), 'h1':clean(h1.group(1) if h1 else ''), 'body_bytes':len(body.encode()), 'server_503_signal':signal, 'classification':cls})
        print(st, cls, url)
print(f'WROTE {out} ({len(urls)} URLs)')
